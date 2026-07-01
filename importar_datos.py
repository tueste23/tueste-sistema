"""
Script de importación de datos TUESTE desde Excel al sistema.
Importa: ingredientes, productos y recetas.
"""
import openpyxl
import requests
import json

BASE_URL = "https://web-production-47696.up.railway.app"
EMAIL = "admin@tueste.com"
PASSWORD = "tueste2024"

# ─── LOGIN ────────────────────────────────────────────────────────────────────
print("🔐 Iniciando sesión...")
r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": EMAIL, "password": PASSWORD})
if "access_token" not in r.json():
    print("Error login:", r.text)
    exit(1)
token = r.json()["access_token"]
headers = {"Authorization": f"Bearer {token}"}
print("✅ Sesión iniciada")

# ─── CARGAR EXCEL ─────────────────────────────────────────────────────────────
print("\n📂 Cargando Excel...")
wb = openpyxl.load_workbook("RECETAS MASTER.xlsx", data_only=True)

# ─── 1. INGREDIENTES desde PRECIO PROVEEDORES ─────────────────────────────────
print("\n🧂 Importando ingredientes...")
ws_precios = wb["PRECIO PROVEEDORES"]
ingredientes_excel = {}  # nombre_lower -> {nombre, precio, unidad}

for i, row in enumerate(ws_precios.iter_rows(values_only=True)):
    if i <= 1:
        continue
    nombre = row[0]
    precio = row[1]
    unidad = row[2]
    if not nombre or nombre == "LISTADO DE PRODUCTOS":
        continue
    nombre = str(nombre).strip()
    try:
        precio_float = float(precio) if precio else 0.0
    except:
        precio_float = 0.0
    unidad_str = str(unidad).strip() if unidad else "kg"
    ingredientes_excel[nombre.lower()] = {
        "nombre": nombre,
        "precio": precio_float,
        "unidad": unidad_str
    }

print(f"  → {len(ingredientes_excel)} ingredientes en Excel")

# Subir ingredientes
ingrediente_id_map = {}  # nombre_lower -> id en sistema
ok = 0
fail = 0
for key, data in ingredientes_excel.items():
    body = {
        "nombre": data["nombre"],
        "unidad": data["unidad"],
        "costo_actual": data["precio"],
        "stock_actual": 0.0,
        "stock_minimo": 0.0
    }
    r = requests.post(f"{BASE_URL}/api/ingredientes", json=body, headers=headers)
    if r.status_code in (200, 201):
        ing_id = r.json()["id"]
        ingrediente_id_map[key] = ing_id
        ok += 1
    else:
        fail += 1
        if fail <= 3:
            print(f"  ⚠ Error en '{data['nombre']}': {r.text[:100]}")

print(f"  ✅ {ok} ingredientes importados, {fail} errores")

# Traer lista actualizada del sistema (por si ya existían)
r = requests.get(f"{BASE_URL}/api/ingredientes", headers=headers)
for ing in r.json():
    ingrediente_id_map[ing["nombre"].lower()] = ing["id"]

# ─── 2. PRODUCTOS desde TUESTE AL PÚBLICO ─────────────────────────────────────
print("\n🍽️ Importando productos...")
ws_pub = wb["TUESTE AL PÚBLICO"]
productos_excel = {}  # nombre -> precio_venta

for i, row in enumerate(ws_pub.iter_rows(values_only=True)):
    if i < 2:
        continue
    nombre = row[1]
    precio_venta = row[4]
    if not nombre or not precio_venta:
        continue
    nombre = str(nombre).strip()
    try:
        pv = float(precio_venta)
    except:
        pv = 0.0
    productos_excel[nombre] = pv

print(f"  → {len(productos_excel)} productos en Excel")

producto_id_map = {}  # nombre_lower -> id
ok = 0
for nombre, precio in productos_excel.items():
    body = {"nombre": nombre, "precio_venta": precio}
    r = requests.post(f"{BASE_URL}/api/productos", json=body, headers=headers)
    if r.status_code in (200, 201):
        producto_id_map[nombre.lower()] = r.json()["id"]
        ok += 1

print(f"  ✅ {ok} productos importados")

# ─── 3. RECETAS desde COCINA TUESTE y CAFETERIA TUESTE ────────────────────────
def extraer_recetas_horizontal(ws):
    """Extrae recetas del formato horizontal del Excel."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    row_nombres = rows[1]  # fila 2: nombres de recetas

    # Encontrar columnas donde arrancan recetas
    recipe_cols = []
    for i, val in enumerate(row_nombres):
        if val and str(val).upper().strip() not in ["KG/LT/UNI", "TOTAL $", "$", ""]:
            recipe_cols.append((i, str(val).strip()))

    recetas = []
    for col_idx, nombre_receta in recipe_cols:
        # Ingredientes: col_idx = nombre, col_idx+1 = cantidad
        items = []
        for row in rows[2:]:  # desde fila 3 en adelante
            ingrediente_nombre = row[col_idx] if col_idx < len(row) else None
            cantidad = row[col_idx + 1] if (col_idx + 1) < len(row) else None

            if not ingrediente_nombre or not cantidad:
                continue
            ingrediente_nombre = str(ingrediente_nombre).strip()
            if ingrediente_nombre.lower() in ["kg/lt/uni", "rinde", "porcion", ""]:
                continue
            try:
                qty = float(cantidad)
                if qty > 0:
                    items.append({"ingrediente": ingrediente_nombre, "cantidad": qty})
            except:
                pass

        if items:
            recetas.append({"nombre": nombre_receta, "items": items})

    return recetas

print("\n📋 Importando recetas...")

# Primero crear los productos de cocina/cafetería que no estén en TUESTE AL PÚBLICO
ws_cocina = wb["COCINA TUESTE"]
ws_cafe = wb["CAFETERIA TUESTE"]
ws_subrecetas = wb["SUBRECETAS TUESTE  CP"]

recetas_cocina = extraer_recetas_horizontal(ws_cocina)
recetas_cafe = extraer_recetas_horizontal(ws_cafe)
recetas_sub = extraer_recetas_horizontal(ws_subrecetas)

todas_recetas = recetas_cocina + recetas_cafe + recetas_sub

print(f"  → {len(recetas_cocina)} recetas de Cocina, {len(recetas_cafe)} de Cafetería, {len(recetas_sub)} subrecetas")

# Crear productos faltantes
for receta in todas_recetas:
    nombre_lower = receta["nombre"].lower()
    if nombre_lower not in producto_id_map:
        body = {"nombre": receta["nombre"], "precio_venta": 0.0}
        r = requests.post(f"{BASE_URL}/api/productos", json=body, headers=headers)
        if r.status_code in (200, 201):
            producto_id_map[nombre_lower] = r.json()["id"]

# Refrescar mapa de productos
r = requests.get(f"{BASE_URL}/api/productos", headers=headers)
for p in r.json():
    producto_id_map[p["nombre"].lower()] = p["id"]

# Subir recetas
ok_recetas = 0
fail_recetas = 0
no_match = []

for receta in todas_recetas:
    nombre_lower = receta["nombre"].lower()
    producto_id = producto_id_map.get(nombre_lower)
    if not producto_id:
        fail_recetas += 1
        continue

    items_payload = []
    for item in receta["items"]:
        ing_nombre_lower = item["ingrediente"].lower()
        ing_id = ingrediente_id_map.get(ing_nombre_lower)

        # Búsqueda aproximada si no matchea exacto
        if not ing_id:
            for key in ingrediente_id_map:
                if ing_nombre_lower in key or key in ing_nombre_lower:
                    ing_id = ingrediente_id_map[key]
                    break

        if ing_id:
            items_payload.append({
                "ingrediente_id": ing_id,
                "cantidad": item["cantidad"]
            })
        else:
            no_match.append(item["ingrediente"])

    if not items_payload:
        fail_recetas += 1
        continue

    body = {
        "producto_id": producto_id,
        "rendimiento": 1.0,
        "items": items_payload
    }
    r = requests.post(f"{BASE_URL}/api/recetas", json=body, headers=headers)
    if r.status_code in (200, 201):
        ok_recetas += 1
    else:
        fail_recetas += 1

print(f"  ✅ {ok_recetas} recetas importadas, {fail_recetas} fallidas")

if no_match:
    no_match_uniq = list(set(no_match))[:10]
    print(f"  ⚠ Ingredientes sin match ({len(set(no_match))} únicos): {no_match_uniq}")

print("\n🎉 ¡Importación completa!")
print(f"   Ingredientes: {len(ingrediente_id_map)}")
print(f"   Productos: {len(producto_id_map)}")
print(f"   Recetas: {ok_recetas}")
